import pandas as pd
import openpyxl as xl
import tkinter as tk
from tkinter import filedialog


def select_file(init_dir=None, title=None):
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(initialdir=init_dir, title=title)
    root.destroy()
    return filepath

def key_plus_df(df, rows, iter, start, col=0):
    end = rows[iter+1] if iter+1 < len(rows) else None
    key = df.loc[start,col].strip()
    df = df.loc[start:end-1] if end is not None else df.loc[start:]
    return key, df

def process_spreadsheet(filepath):
    df = pd.read_excel(filepath, header=None)
    
    # fill missing values with empty string
    df.fillna('', inplace=True)
    
    # get row names that have 'Carrier' in them
    for col in df.columns:
        if df[col].astype(str).str.contains('Carrier:').any():
            break
    carrier_rows = df[df.loc[:,col].str.contains('Carrier:')].index.to_list()

    # create a dictionary of carriers
    charts = {}
    
    for i, car_start in enumerate(carrier_rows):
        # get the chart row names for the current carrier
        carrier, car_df = key_plus_df(df, carrier_rows, i, car_start, col)
        print("Working with", carrier, "...")

        chart_rows = car_df[car_df.loc[:,col+1].str.contains('Chart #')].index.to_list()
        charts[carrier] = {"Carrier info": df.loc[car_start:chart_rows[0]-1]}

        for j, ch_start in enumerate(chart_rows):
            chart, ch_df = key_plus_df(car_df, chart_rows, j, ch_start,col+1)
            print(5*"\t", chart,"...")
            
            visit_rows = ch_df[ch_df.iloc[:,col+1].str.contains('Service Date')].index.to_list()
            charts[carrier][chart] = {"Chart info": ch_df.loc[ch_start:visit_rows[0]-1]}

            for k, v_start in enumerate(visit_rows):
                date, v_df = key_plus_df(ch_df, visit_rows, k, v_start,col)
                print(7*"\t", date)

                charts[carrier][chart][date] = v_df.reset_index(drop=True)
        
    return charts



print('Select the new spreadsheet')
new = select_file(title = 'Select the new spreadsheet')
print('Select the old spreadsheet')
old = select_file(title='Select the old spreadsheet')

new_charts = process_spreadsheet(new)
old_charts = process_spreadsheet(old)

df = pd.DataFrame(columns = ["state"])
for carrier in new_charts.keys():
    if carrier not in old_charts.keys():
        print(f'New "{carrier}"')
        df = pd.concat([df,new_charts[carrier]['Carrier info']], ignore_index=True)
        car_list = [chart for key,chart in new_charts[carrier].items() if 'Carrier info' not in key]
        for chart in car_list:
            char_df = pd.concat(list(chart.values()),ignore_index=True)
            df = pd.concat([df,char_df],ignore_index=True)
        df["state"].fillna('new', inplace=True)
         
    else:
        charts  = set(new_charts[carrier].keys()) | set(old_charts[carrier].keys())
        charts.discard('Carrier info')
        car_df = new_charts[carrier]['Carrier info']
        df = pd.concat([df,car_df])
        df["state"].fillna('old', inplace=True)

        for chart in charts:
            if chart not in old_charts[carrier].keys():
                print(f'New {chart} for "{carrier}"')
                chart_df = pd.concat(list(new_charts[carrier][chart].values()),ignore_index=True)
                df = pd.concat([df,chart_df])
                df["state"].fillna('new', inplace=True)

            elif chart not in new_charts[carrier].keys():
                print(f'{chart} for "{carrier}" was removed')
                chart_df = pd.concat(list(old_charts[carrier][chart].values()),ignore_index=True)
                df = pd.concat([df,chart_df])
                df["state"].fillna('removed', inplace=True)

            else:
                df = pd.concat([df,new_charts[carrier][chart]['Chart info']],ignore_index=True)
                df["state"].fillna('old', inplace=True)
                dates = set(new_charts[carrier][chart].keys()) | set(old_charts[carrier][chart].keys())
                dates.discard('Chart info')
                
                for date in dates:
                    if date not in old_charts[carrier][chart].keys():
                        df = pd.concat([df,new_charts[carrier][chart][date]],ignore_index=True)
                        df["state"].fillna('new', inplace=True)
                    elif date not in new_charts[carrier][chart].keys():
                        df = pd.concat([df,old_charts[carrier][chart][date]],ignore_index=True)
                        df["state"].fillna('removed', inplace=True)
                    else:
                        df = pd.concat([df,new_charts[carrier][chart][date]],ignore_index=True)
                        df["state"].fillna('old', inplace=True)
            df.loc[len(df)] = ''
                

new_filepath = 'Spreadsheet_Updated.xlsx'
sheet = 'CarrierArDetail'

df.to_excel(new_filepath, sheet_name=sheet, startrow=0, header=False, index=False)

wb = xl.load_workbook(new_filepath)
ws = wb[sheet]

font_base = xl.styles.Font(name='Tahoma', size=7, bold=False)
font_bold = xl.styles.Font(name='Tahoma', size=7, bold=True)
font_car = xl.styles.Font(name='Tahoma', size=8, bold=True)
bold_text = ["Chart #:", "Patient Name:","Date of birth:", "Aging Date:", "Provider", "CPT Code","Service Date", "Total"]
money_fmt = '$#,##0.00'
percent_fmt = '0.00%'
car_row_prev = False

for row in ws.iter_rows(min_row=1, max_row=df.shape[0]):
    if row[0].value == 'new':
        col = 'FFFF00'
        typ = 'solid'
    elif row[0].value == 'removed':
        col = 'FF0000'
        typ = 'solid'
    else:
        col = None
        typ = None
    
    for cell in row:
        # define the fill color
        fill_col = xl.styles.PatternFill(start_color=col, end_color=col, fill_type=typ)
        cell.fill = fill_col
        # define the font
        if type(cell.value) is str and ("Carrier: " in cell.value or "Phone:" in cell.value):
            cell.font = font_car
        elif type(cell.value) is str and any([val in cell.value for val in bold_text]):
            cell.font = font_bold
        elif cell.column in range(7,14):
            cell.font = font_bold
        else:
            cell.font = font_base
        # define the number format
        if car_row_prev:
            cell.number_format = percent_fmt
        elif cell.column > 6:
            cell.number_format = money_fmt
    car_row_prev = True if type(row[1].value) is str and "Carrier: " in row[1].value else False

wb.save(new_filepath)